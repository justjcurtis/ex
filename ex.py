import jpyHelpers
import Excel
import openpyxl
from tabulate import tabulate

#globals
available_commands = ["display"]

def main():
    global book
    global available_commands

    running = True
    while(running):
        jpyHelpers.clear()
        print("drop excel sheet here")
        print("or type exit to exit")
        path = input("=> ")
        if(path.lower() == "x" or path.lower() == "exit"):
            jpyHelpers.clear()
            running = False
            continue
        book = openpyxl.load_workbook(path)

        while(True):
            menu = createBookMenu(book)
            command = jpyHelpers.displayMenu(menu, available_commands)
            if(command == "back"):
                break
            else:
                proccessCommand(command)


#Menu

def proccessCommand(command):
    print()
    if(command[0].lower() == "display" or command[0].lower() == "d"):
        jpyHelpers.clear()
        print()
        print("Displaying " + book.sheetnames[command[1]] + " ...")
        table = getTable(book[book.sheetnames[command[1]]])
        table = trimTable(table)
        printTable(table, table[1])
    else:
        print("command not found")
        print("command : " + command[0])
    print()
    input("press enter to continue")

def createBookMenu(book):
    menu = [["Index", "Sheet Name"]]
    for i in range(len(book.sheetnames)):
        menu.append([i+1, book.sheetnames[i]])
    return menu

#Methods

def printTable(table, _headers = None):
    table = list(table)
    if(_headers == None):
        print(tabulate(table, table[1]))
    else:
        print(tabulate(stripTableRows(table, _headers), _headers))

def getSheetByIndex(book, i):
    return book[book.sheetnames[i]]

def stripTableRows(table, row):
    row = list(row)
    table = list(table)

    stripping = True

    while stripping:
        stripping = False
        for i in range(len(table)):
            if(table[i] == row):
                stripping = True
                del table[i]
                break
    return table

def trimTable(table):

    table = list(table)

    for row in table:
        #backsearch
        maxRowIndex = len(row) - 1
        if(maxRowIndex > -1):
            for i in range(maxRowIndex):
                if(row[maxRowIndex-i] == ''):
                    del row[maxRowIndex-i]
                else:
                    break

        #frontsearch
        maxRowIndex = len(row) - 1
        if(maxRowIndex > -1):
            for i in range(maxRowIndex):
                if(row[i] == ''):
                    del row[i]
                else:
                    break

    #strip empty rows
    table = stripTableRows(table, [''])

    return table


def getTable(sheet):

    record = []
    table = []
    for row in sheet:
        for cell in row:
            if not (cell.value == None):
                record.append(cell.value)
            else:
                record.append('')
        table.append(record)
        record = []

    return table
    
def replaceEmptyCells(sheet):

    for row in sheet:
        for cell in row:
            if(cell.value == None):
                cell.value = "EMPTY"

    return sheet

main()
